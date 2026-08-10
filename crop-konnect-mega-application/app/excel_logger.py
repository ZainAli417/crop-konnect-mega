import logging
import threading
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.config import Settings
from app.schemas import ReadingResponse


logger = logging.getLogger(__name__)

_LOCK = threading.Lock()
_BASE_DIR = Path(__file__).resolve().parent.parent
_HEADERS = [
    "Date",
    "Day",
    "Time",
    "Reading ID",
    "Device ID",
    "Station Name",
    "Recorded At",
    "Received At",
    "Wind Speed (m/s)",
    "Wind Direction (deg)",
    "Wind Direction",
    "Soil Moisture (%)",
    "Soil Temp (C)",
    "Soil EC (uS/cm)",
    "Nitrogen (mg/kg)",
    "Phosphorus (mg/kg)",
    "Potassium (mg/kg)",
    "Soil pH",
    "Rainfall (mm)",
    "Solar Radiation (W/m2)",
]


def ensure_today_excel_log(settings: Settings) -> Path | None:
    if not settings.excel_logging_enabled:
        return None

    local_now = datetime.now(_timezone(settings.excel_log_timezone))
    path = _log_path(settings, local_now)
    with _LOCK:
        _ensure_workbook(path)
    return path


def append_reading_to_excel_log(reading: ReadingResponse, settings: Settings) -> Path | None:
    if not settings.excel_logging_enabled:
        return None

    local_recorded_at = _to_local(reading.recorded_at, settings.excel_log_timezone)
    local_received_at = _to_local(reading.received_at, settings.excel_log_timezone)
    path = _log_path(settings, local_recorded_at)

    row = [
        local_recorded_at.date(),
        local_recorded_at.strftime("%A"),
        local_recorded_at.time().replace(microsecond=0),
        reading.id,
        reading.device_id,
        reading.station_name,
        local_recorded_at.replace(tzinfo=None),
        local_received_at.replace(tzinfo=None),
        reading.ws,
        reading.wd_deg,
        reading.wd_dir,
        reading.moist,
        reading.temp,
        reading.ec,
        reading.n,
        reading.p,
        reading.k,
        reading.ph,
        reading.rain,
        reading.solar,
    ]

    with _LOCK:
        workbook, sheet = _ensure_workbook(path)
        sheet.append(row)
        _format_sheet(sheet)
        workbook.save(path)

    return path


def _timezone(name: str) -> ZoneInfo:
    try:
        return ZoneInfo(name)
    except ZoneInfoNotFoundError:
        logger.warning("Unknown Excel log timezone %s, falling back to UTC", name)
        return ZoneInfo("UTC")


def _to_local(value: datetime, timezone_name: str) -> datetime:
    source = value
    if source.tzinfo is None:
        source = source.replace(tzinfo=timezone.utc)
    return source.astimezone(_timezone(timezone_name))


def _log_dir(settings: Settings) -> Path:
    configured = Path(settings.excel_log_dir).expanduser()
    if configured.is_absolute():
        return configured
    return _BASE_DIR / configured


def _log_path(settings: Settings, local_datetime: datetime) -> Path:
    date_text = local_datetime.strftime("%Y-%m-%d")
    day_text = local_datetime.strftime("%A")
    return _log_dir(settings) / f"sensor_readings_{date_text}_{day_text}.xlsx"


def _ensure_workbook(path: Path) -> tuple[Workbook, Worksheet]:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        workbook = load_workbook(path)
        sheet = workbook.active
        if sheet.max_row == 0 or sheet.cell(row=1, column=1).value != _HEADERS[0]:
            _write_headers(sheet)
            workbook.save(path)
        return workbook, sheet

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Readings"
    _write_headers(sheet)
    _format_sheet(sheet)
    workbook.save(path)
    return workbook, sheet


def _write_headers(sheet: Worksheet) -> None:
    for column_index, header in enumerate(_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)


def _format_sheet(sheet: Worksheet) -> None:
    widths = {
        "A": 12,
        "B": 12,
        "C": 10,
        "D": 12,
        "E": 16,
        "F": 20,
        "G": 20,
        "H": 20,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for column_index in range(9, len(_HEADERS) + 1):
        sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = 18
    sheet.freeze_panes = "A2"
